from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils import get_column_letter

OUTPUT_FILE = "Metal_Sector_Dashboard.xlsx"

thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="DDDDDD")
button_fill = PatternFill("solid", fgColor="CFE2F3")


def auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value) if cell.value is not None else ""
            except Exception:
                value = ""
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def style_headers(ws, header_row=1):
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        cell.fill = header_fill


def add_table(ws, start_row, headers, data):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
    style_headers(ws, header_row=start_row)
    for r, row in enumerate(data, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
    auto_width(ws)


def add_home_sheet(wb):
    ws = wb.create_sheet("Home")
    ws.sheet_view.showGridLines = True
    title = ws.cell(row=1, column=1, value="Metal Sector Dashboard")
    title.font = Font(size=18, bold=True)
    title.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Links to companies
    ws.cell(row=3, column=1, value="Tata Steel").font = Font(bold=True)
    ws.cell(row=3, column=1).hyperlink = "#TataSteel!A1"
    ws.cell(row=3, column=1).style = "Hyperlink"

    ws.cell(row=4, column=1, value="JSW Steel").font = Font(bold=True)
    ws.cell(row=4, column=1).hyperlink = "#JSWSteel!A1"
    ws.cell(row=4, column=1).style = "Hyperlink"

    auto_width(ws)
    return ws


def add_company_sheet(wb, sheet_name, display_name, partner_prefix):
    ws = wb.create_sheet(sheet_name)
    title = ws.cell(row=1, column=1, value=display_name)
    title.font = Font(size=14, bold=True)
    title.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value="CEO:")
    ws.cell(row=2, column=2, value="[Placeholder]")
    ws.cell(row=3, column=1, value="CFO:")
    ws.cell(row=3, column=2, value="[Placeholder]")

    # Buttons (formatted cells with hyperlinks)
    buttons = [
        ("Share Price", f"#{partner_prefix}_SharePrice!A1"),
        ("Balance Sheet", f"#{partner_prefix}_BalanceSheet!A1"),
        ("Cash Flow", f"#{partner_prefix}_CashFlow!A1"),
        ("Financial Ratios", f"#{partner_prefix}_Ratio!A1"),
    ]
    row = 5
    col = 1
    for text, link in buttons:
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = button_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.hyperlink = link
        cell.style = "Hyperlink"
        ws.row_dimensions[row].height = 18
        col += 2

    # Back to Home
    back = ws.cell(row=10, column=1, value="Back to Home")
    back.hyperlink = "#Home!A1"
    back.style = "Hyperlink"

    auto_width(ws)
    return ws


def create_shareprice_sheet(wb, sheet_name, sample_prices):
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=sheet_name.replace('_', ' '))
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=2, column=1, value="Back")
    # We'll link back programmatically later by caller
    headers = ["Date", "Open", "High", "Low", "Close", "Volume"]
    add_table(ws, start_row=4, headers=headers, data=sample_prices)

    # Chart - Line chart using Close over dates
    chart = LineChart()
    chart.title = "Close Price"
    data_ref = Reference(ws, min_col=5, min_row=4, max_row=4 + len(sample_prices))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sample_prices))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H4")
    auto_width(ws)
    return ws


def create_balance_sheet_sheet(wb, sheet_name, sample_bs):
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=sheet_name.replace('_', ' ')).font = Font(bold=True)
    headers = ["Year", "Assets", "Liabilities", "Equity"]
    add_table(ws, start_row=4, headers=headers, data=sample_bs)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Balance Sheet"
    data = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=4 + len(sample_bs))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sample_bs))
    chart.add_data(data, from_rows=False, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H4")
    auto_width(ws)
    return ws


def create_cashflow_sheet(wb, sheet_name, sample_cf):
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=sheet_name.replace('_', ' ')).font = Font(bold=True)
    headers = ["Year", "Operating", "Investing", "Financing"]
    add_table(ws, start_row=4, headers=headers, data=sample_cf)
    chart = LineChart()
    chart.title = "Cash Flow Components"
    data = Reference(ws, min_col=2, min_row=4, max_col=4, max_row=4 + len(sample_cf))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sample_cf))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H4")
    auto_width(ws)
    return ws


def create_ratio_sheet(wb, sheet_name, sample_ratios):
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=sheet_name.replace('_', ' ')).font = Font(bold=True)
    headers = ["Year", "PE Ratio", "ROE", "Debt/Equity", "Current Ratio"]
    add_table(ws, start_row=4, headers=headers, data=sample_ratios)
    chart = LineChart()
    chart.title = "Key Ratios"
    data = Reference(ws, min_col=2, min_row=4, max_col=5, max_row=4 + len(sample_ratios))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sample_ratios))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H4")
    auto_width(ws)
    return ws


def link_back(ws, target_sheet_name):
    # Place a Back link at A2
    cell = ws.cell(row=2, column=1, value=f"Back to {target_sheet_name}")
    cell.hyperlink = f"#{target_sheet_name}!A1"
    cell.style = "Hyperlink"


def build_dashboard():
    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    add_home_sheet(wb)

    # Company sheets
    add_company_sheet(wb, "TataSteel", "Tata Steel", "Tata")
    add_company_sheet(wb, "JSWSteel", "JSW Steel", "JSW")

    # Sample data (small sets)
    tata_prices = [
        ("2025-01-01", 800, 820, 790, 810, 1200000),
        ("2025-01-02", 810, 830, 800, 825, 1100000),
        ("2025-01-03", 825, 840, 820, 835, 900000),
        ("2025-01-04", 835, 850, 830, 845, 950000),
    ]
    jsw_prices = [
        ("2025-01-01", 700, 720, 690, 710, 850000),
        ("2025-01-02", 710, 730, 705, 725, 920000),
        ("2025-01-03", 725, 740, 720, 735, 780000),
        ("2025-01-04", 735, 750, 730, 745, 800000),
    ]

    tata_bs = [
        (2022, 50000, 30000, 20000),
        (2023, 55000, 32000, 23000),
        (2024, 60000, 35000, 25000),
    ]
    jsw_bs = [
        (2022, 40000, 22000, 18000),
        (2023, 45000, 24000, 21000),
        (2024, 48000, 26000, 22000),
    ]

    tata_cf = [
        (2022, 3000, -500, -1000),
        (2023, 3500, -600, -1100),
        (2024, 3700, -550, -1200),
    ]
    jsw_cf = [
        (2022, 2500, -400, -900),
        (2023, 2800, -450, -950),
        (2024, 3000, -480, -980),
    ]

    tata_ratios = [
        (2022, 12.5, 0.15, 0.5, 1.3),
        (2023, 13.1, 0.16, 0.48, 1.35),
        (2024, 14.0, 0.17, 0.45, 1.4),
    ]
    jsw_ratios = [
        (2022, 10.5, 0.12, 0.6, 1.2),
        (2023, 11.2, 0.13, 0.58, 1.25),
        (2024, 11.8, 0.14, 0.55, 1.3),
    ]

    # Create data sheets for Tata
    t_sp = create_shareprice_sheet(wb, "Tata_SharePrice", tata_prices)
    link_back(t_sp, "TataSteel")
    t_bs = create_balance_sheet_sheet(wb, "Tata_BalanceSheet", tata_bs)
    link_back(t_bs, "TataSteel")
    t_cf = create_cashflow_sheet(wb, "Tata_CashFlow", tata_cf)
    link_back(t_cf, "TataSteel")
    t_rat = create_ratio_sheet(wb, "Tata_Ratio", tata_ratios)
    link_back(t_rat, "TataSteel")

    # Create data sheets for JSW
    j_sp = create_shareprice_sheet(wb, "JSW_SharePrice", jsw_prices)
    link_back(j_sp, "JSWSteel")
    j_bs = create_balance_sheet_sheet(wb, "JSW_BalanceSheet", jsw_bs)
    link_back(j_bs, "JSWSteel")
    j_cf = create_cashflow_sheet(wb, "JSW_CashFlow", jsw_cf)
    link_back(j_cf, "JSWSteel")
    j_rat = create_ratio_sheet(wb, "JSW_Ratio", jsw_ratios)
    link_back(j_rat, "JSWSteel")

    # Add Back to Home links on company sheets (ensure present)
    wb["TataSteel"].cell(row=10, column=1).hyperlink = "#Home!A1"
    wb["JSWSteel"].cell(row=10, column=1).hyperlink = "#Home!A1"

    wb.save(OUTPUT_FILE)
    print(f"Saved workbook to {OUTPUT_FILE}")


if __name__ == "__main__":
    build_dashboard()
